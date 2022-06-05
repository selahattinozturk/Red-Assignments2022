from openpyxl import Workbook, load_workbook

# OPENING THE FILE TO IMPORT DATAS
wb1 = load_workbook(
    '4- BOBİ FRS Nakit Akış Tablosu - Dolaylı Yöntem (Konsolide).xlsx')
sheets = wb1.sheetnames
ws1 = wb1[sheets[0]]

# CREATING EXCEL FILE
wbNat = Workbook()
wsNat = wbNat.active
wsNat.title = 'BOBİ FRS NAT Dolaylı Konsolide'
wsNat['A1'] = "ACIKLAMALAR"
wsNat['B1'] = "CARI_DONEM"
wsNat['C1'] = "GECMIS_DONEM"


# DATAS ARE IN C,D AND E COLUMNS; CARI_DONEM = K, GECMIS_DONEM = L\
# APPENDING THE VALUES
for i in range(2, 73):
    if(ws1['B' + str(i)].value != None):
        wsNat.append([ws1['B' + str(i)].value, 0, 0])
    elif(ws1['C' + str(i)].value != None):
        wsNat.append([ws1['C' + str(i)].value, 0, 0])
    elif(ws1['D' + str(i)].value != None):
        wsNat.append([ws1['D' + str(i)].value, 0, 0])
    elif(ws1['E' + str(i)].value != None):
        wsNat.append([ws1['E' + str(i)].value, 0, 0])


wbNat.save("4- BOBİ FRS Nakit Akış Tablosu.xlsx")
